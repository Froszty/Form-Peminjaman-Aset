#========= GUNAWAN LUMBANTOBING - SEPTEMBER 2025 =========#

import os
import io
import re
import json
import base64
import sqlite3
from datetime import datetime, timedelta
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, g, abort, send_file
)

# ========= Konfigurasi LDAP / AD =========
LDAP_SERVER = os.getenv("LDAP_SERVER", "ldap://192.168.1.162")
LDAP_DOMAIN = os.getenv("LDAP_DOMAIN", "yakult")            # NetBIOS / short domain (tanpa .local)
LDAP_BASE_DN = os.getenv("LDAP_BASE_DN", "DC=yakult,DC=local")
IT_GROUP_KEYWORDS = [s.strip().lower() for s in os.getenv("IT_GROUP_KEYWORDS", "IT").split(",")]

# ========= App config =========
APP_SECRET = os.environ.get("APP_SECRET", "super-secret-key")
DB_PATH = os.environ.get("DB_PATH", "loans.db")

# ========= Excel template path =========
DEFAULT_TEMPLATE_XLSX = os.path.join(os.path.dirname(__file__), "Form_Peminjaman.xlsx")
TEMPLATE_XLSX_PATH = os.environ.get("TEMPLATE_XLSX_PATH", DEFAULT_TEMPLATE_XLSX)

# --- Ukuran tanda tangan pada export Excel (bisa via ENV) ---
SIG_W_CM = float(os.getenv("SIG_W_CM", "2.5"))
SIG_H_CM = float(os.getenv("SIG_H_CM", "1.8"))
SIG_CROP = os.getenv("SIG_CROP", "0") == "1"   # 0=fit (padding), 1=crop center

# ========= Daftar device seed =========
DEVICE_CODES = [
    "MICROPHONE SAMSON","WEBCAM LOGITECH","CONVERTER VGA TO HDMI","FAC-SKB-023","FAC30","PST-IAU-003","PST-IAU-005","PST-IAU-006","PST-ITA-026",
    "PST-PGA-007","PST-PGA-008","YIP-MGR-038","YIP-MGR-040","YIP-MGR-046","YIP-MGR-047",
    "YIP-MGR-049","YIP-MGR-050","YIP-MGR-051","YIP-MGR-053","YIP-MGR-054","YIP-MGR-055",
    "YIP-MGR-056","YIP-MGR-057","YIP-MGR-058","YIP-MGR-064","YIP-MGR-068","YIP-MGR-083",
    "YIP-MGR-084","YIP-MGR-093","YIP-MGR-114","YIP-SPV-072","YIP-SPV-075","YIP-SPV-078",
    "YIP-SPV-079","YIP-SPV-081","YIP-SPV-083","YIP-SPV-084","YIP-SPV-085","YIP-SPV-087",
    "YIP-SPV-090","YIP-SPV-091","YIP-SPV-093","YIP-SPV-094","YIP-SPV-234","YIP-SPV-235",
    "YIP-SPV-240","PROJECTOR PANASONIC"
]

# ========= Daftar Divisi (untuk pilihan manual user non-IT) =========
DIVISIONS = [
    "Accounting", "Advisor", "AR DS", "AR YL",
    "Billing", "Direct Sales", "Fac. Mojokerto", "HRD GA", "Internal Audit", "IT", "Legal", "Management Factory", "Management Finance",
    "Marcomm", "MOS", "Procurement", "Sekretariat", "SPV Cabang", "Tax", "TND",
    "Treasury", "Yakult Lady", "OB"
]

# ========= Flask init =========
app = Flask(__name__)
app.secret_key = APP_SECRET
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    TEMPLATES_AUTO_RELOAD=True,
)
app.jinja_env.auto_reload = True

# ========= LDAP (ldap3) =========
from ldap3 import Server, Connection, ALL, ALL_ATTRIBUTES, SUBTREE, core

def _user_upn(username: str) -> str:
    # user -> user@yakult.local
    return f"{username}@{LDAP_DOMAIN}.local"

def _user_netbios(username: str) -> str:
    # user -> YAKULT\user
    return f"{LDAP_DOMAIN}\\{username}"

def _cn_from_dn(dn: str) -> str:
    # "CN=IT,OU=Groups,DC=yakult,DC=local" -> "IT"
    try:
        parts = [p for p in dn.split(",") if p.strip().upper().startswith("CN=")]
        if parts:
            return parts[0].split("=", 1)[1]
    except Exception:
        pass
    return dn

def _is_it_group(groups_cn_lower):
    for cn in groups_cn_lower:
        for key in IT_GROUP_KEYWORDS:
            if key in cn:
                return True
    return False

def ad_authenticate(username: str, password: str):
    """
    Bind ke AD pakai kredensial user.
    Ambil: displayName + memberOf (untuk deteksi IT).
    NOTE: hanya baca atribut dasar (belum pakai service bind).
    """
    if not username or not password:
        return None
    server = Server(LDAP_SERVER, get_info=ALL)
    tries = [_user_upn(username), _user_netbios(username), username]
    for user_id in tries:
        try:
            conn = Connection(server, user=user_id, password=password, auto_bind=True)
            filt = f"(&(objectClass=user)(sAMAccountName={username}))"
            ok = conn.search(LDAP_BASE_DN, filt, SUBTREE, attributes=ALL_ATTRIBUTES)
            if not ok or not conn.entries:
                upn = _user_upn(username)
                filt2 = f"(&(objectClass=user)(userPrincipalName={upn}))"
                conn.search(LDAP_BASE_DN, filt2, SUBTREE, attributes=ALL_ATTRIBUTES)

            if not conn.entries:
                conn.unbind()
                continue

            entry = conn.entries[0]
            display_name = str(entry.displayName) if "displayName" in entry else username
            groups_dn = [str(x) for x in entry.memberOf] if ("memberOf" in entry and entry.memberOf) else []
            groups_cn = [_cn_from_dn(dn) for dn in groups_dn]
            groups_cn_lower = [cn.lower() for cn in groups_cn]
            is_it = _is_it_group(groups_cn_lower)
            conn.unbind()
            return {
                "username": username,
                "display_name": display_name,
                "groups_cn": groups_cn,
                "groups_cn_lower": groups_cn_lower,
                "is_it": is_it
            }
        except core.exceptions.LDAPBindError:
            pass
        except Exception:
            pass
    return None

# ========= DB Helpers =========
from flask import g

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(_exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()

def column_exists(db, table, col):
    rows = db.execute(f"PRAGMA table_info({table})").fetchall()
    return any(r["name"] == col for r in rows)

def init_db():
    db = get_db()

    # devices
    db.execute("""
        CREATE TABLE IF NOT EXISTS devices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE,
            status TEXT  -- available | reserved | allocated
        )
    """)

    # loans
    db.execute("""
        CREATE TABLE IF NOT EXISTS loans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT,
            start_date TEXT,
            duration_days INTEGER,
            device_code TEXT,

            borrower_name TEXT,
            borrower_division TEXT,
            borrower_signature TEXT,
            borrower_signed_at TEXT,

            status TEXT,  -- submitted | active | return_submitted | returned
            status_kondisi TEXT, -- Baik | Rusak
            keterangan_kondisi TEXT,

            it_confirmed_by TEXT,
            it_signature TEXT,
            it_confirmed_at TEXT,

            return_borrower_name TEXT,
            return_division TEXT,
            return_signature TEXT,
            return_submitted_at TEXT,

            return_it_confirmed_by TEXT,
            return_it_checked_box INTEGER,
            return_it_confirmed_at TEXT,
            return_it_signature TEXT,
            return_status_kondisi TEXT,
            return_keterangan_kondisi TEXT,

            form_seq INTEGER,
            form_month INTEGER,
            form_year INTEGER
        )
    """)

    # users (baru) – catat user AD & divisi terpilih untuk non-IT
    db.execute("""
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            display_name TEXT,
            is_it INTEGER DEFAULT 0,
            division TEXT,               -- pilihan user (non-IT)
            groups_json TEXT,            -- opsional: simpan CN groups
            first_login_at TEXT,
            last_login_at TEXT
        )
    """)

    # back-compat add columns
    for col, type_ in [
        ("device_code", "TEXT"),
        ("return_submitted_at", "TEXT"),
        ("return_it_checked_box", "INTEGER"),
        ("return_it_signature", "TEXT"),
        ("form_seq", "INTEGER"),
        ("form_month", "INTEGER"),
        ("form_year", "INTEGER"),
        ("status_kondisi", "TEXT"),
        ("keterangan_kondisi", "TEXT"),
        ("return_status_kondisi", "TEXT"),
        ("return_keterangan_kondisi", "TEXT"),
    ]:
        if not column_exists(db, "loans", col):
            db.execute(f"ALTER TABLE loans ADD COLUMN {col} {type_}")

    # seed devices
    for code in DEVICE_CODES:
        db.execute("INSERT OR IGNORE INTO devices(code, status) VALUES (?, 'available')", (code,))
    db.commit()

@app.before_request
def _before_request():
    init_db()

# ========= Auth Decorator =========
def login_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if "username" not in session:
            return redirect(url_for("login"))
        return view(**kwargs)
    return wrapped_view

# ========= Filters & Utils =========
def parse_date(date_str):
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(date_str, fmt)
        except Exception:
            pass
    return None

@app.template_filter("remaining_str")
def remaining_str_filter(row):
    if not row["start_date"] or not row["duration_days"]:
        return "-"
    start = parse_date(row["start_date"])
    if not start:
        return "-"
    try:
        dur = int(row["duration_days"])
    except Exception:
        dur = 0
    end = start + timedelta(days=dur)
    now = datetime.now()
    delta = end - now
    if delta.total_seconds() >= 0:
        days = delta.days
        hours = (delta.seconds // 3600)
        mins = (delta.seconds % 3600) // 60
        return f"{days} hari {hours} jam {mins} menit"
    else:
        over = now - end
        days = over.days
        hours = (over.seconds // 3600)
        return f"TERLAMBAT {days} hari {hours} jam"

@app.template_filter("end_date")
def end_date_filter(row):
    if not row["start_date"] or not row["duration_days"]:
        return "-"
    start = parse_date(row["start_date"])
    if not start:
        return "-"
    try:
        dur = int(row["duration_days"])
    except Exception:
        dur = 0
    end = start + timedelta(days=dur)
    return end.strftime("%Y-%m-%d")

# ========= Routes =========
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "").strip()
        info = ad_authenticate(u, p)
        if info:
            # Simpan/Update user ke tabel users
            db = get_db()
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            is_it = 1 if info["is_it"] else 0
            groups_json = json.dumps(info.get("groups_cn", []))

            row = db.execute("SELECT username, division FROM users WHERE username = ?", (info["username"],)).fetchone()
            if row:
                db.execute("""
                    UPDATE users
                    SET display_name=?, is_it=?, groups_json=?, last_login_at=?
                    WHERE username=?
                """, (info["display_name"], is_it, groups_json, now, info["username"]))
            else:
                db.execute("""
                    INSERT INTO users(username, display_name, is_it, division, groups_json, first_login_at, last_login_at)
                    VALUES(?,?,?,?,?,?,?)
                """, (info["username"], info["display_name"], is_it, None, groups_json, now, now))
            db.commit()

            # set session
            session["username"] = info["username"]
            session["display_name"] = info["display_name"]
            session["groups_cn"] = info["groups_cn"]
            session["groups_cn_lower"] = info["groups_cn_lower"]
            session["is_it"] = bool(info["is_it"])

            # ambil division yang pernah dipilih (jika ada)
            my = db.execute("SELECT division FROM users WHERE username=?", (info["username"],)).fetchone()
            user_division = my["division"] if my and my["division"] else None
            session["user_division"] = user_division

            # jika bukan IT dan belum memilih divisi -> redirect pilih divisi
            if not session["is_it"] and not session.get("user_division"):
                return redirect(url_for("select_division"))

            flash("Login berhasil.", "success")
            return redirect(url_for("dashboard"))
        flash("Login gagal. Periksa username/password AD Anda.", "danger")
    return render_template("login.html")

@app.route("/select-division", methods=["GET", "POST"])
@login_required
def select_division():
    # Hanya non-IT yang perlu memilih divisi
    if session.get("is_it"):
        return redirect(url_for("dashboard"))

    db = get_db()
    if request.method == "POST":
        div = request.form.get("division", "").strip()
        if not div:
            flash("Silakan pilih divisi.", "warning")
            return render_template("select_division.html", divisions=DIVISIONS)

        db.execute("UPDATE users SET division=? WHERE username=?", (div, session["username"]))
        db.commit()
        session["user_division"] = div
        flash("Divisi tersimpan.", "success")
        return redirect(url_for("dashboard"))

    return render_template("select_division.html", divisions=DIVISIONS)

@app.route("/logout")
def logout():
    session.clear()
    flash("Anda telah logout.", "info")
    return redirect(url_for("login"))

# Dashboard:
# - IT  : lihat semua peminjaman aktif
# - non-IT: hanya yang borrower_division == division user
@app.route("/")
@login_required
def dashboard():
    db = get_db()
    if session.get("is_it"):
        rows = db.execute(
            "SELECT * FROM loans WHERE status = 'active' ORDER BY created_at DESC"
        ).fetchall()
    else:
        # Pastikan sudah punya division
        if not session.get("user_division"):
            return redirect(url_for("select_division"))
        rows = db.execute(
            "SELECT * FROM loans WHERE status='active' AND lower(borrower_division)=lower(?) ORDER BY created_at DESC",
            (session["user_division"],)
        ).fetchall()
    return render_template("dashboard.html", rows=rows)

# Daftar submitted (untuk IT memproses)
@app.route("/submitted")
@login_required
def submitted_list():
    if not session.get("is_it"):
        abort(403)
    db = get_db()
    rows = db.execute(
        "SELECT * FROM loans WHERE status = 'submitted' ORDER BY created_at DESC"
    ).fetchall()
    return render_template("submitted_list.html", rows=rows)

# Create form peminjaman
@app.route("/create", methods=["GET", "POST"])
@login_required
def create_form():
    db = get_db()
    devices = db.execute(
        "SELECT code FROM devices WHERE status = 'available' ORDER BY code ASC"
    ).fetchall()
    
    # Tentukan divisi peminjam berdasarkan status login
    borrower_division_val = "IT" if session.get("is_it") else (session.get("user_division") or "")

    if request.method == "POST":
        device_code = request.form.get("device_code", "").strip()
        duration_days = request.form.get("duration_days", "0").strip()
        borrower_name = request.form.get("borrower_name", "").strip()
        borrower_division = request.form.get("borrower_division", "").strip()
        signature_data = request.form.get("signature_data", "")

        now_dt = datetime.now()
        now = now_dt.strftime("%Y-%m-%d %H:%M:%S")

        # durasi max 30 hari
        try:
            ddays = int(duration_days)
        except Exception:
            ddays = 0
        if ddays > 30:
            flash("Tidak bisa meminjam lebih dari 30 hari.", "warning")
            devices2 = db.execute("SELECT code FROM devices WHERE status = 'available' ORDER BY code").fetchall()
            return render_template("create_form.html", devices=devices2, server_now=now, borrower_division=borrower_division_val)

        if not all([device_code, duration_days, borrower_name, borrower_division, signature_data]):
            flash("Mohon lengkapi semua field dan tanda tangan.", "warning")
            devices2 = db.execute("SELECT code FROM devices WHERE status = 'available' ORDER BY code").fetchall()
            return render_template("create_form.html", devices=devices2, server_now=now, borrower_division=borrower_division_val)

        drow = db.execute("SELECT * FROM devices WHERE code = ?", (device_code,)).fetchone()
        if not drow or drow["status"] != "available":
            flash("Device tidak tersedia. Refresh daftar device.", "danger")
            return redirect(url_for("create_form"))

        # reset nomor per bulan
        y = now_dt.year
        m = now_dt.month
        seq = db.execute(
            "SELECT COALESCE(MAX(form_seq),0)+1 AS next_seq FROM loans WHERE form_year=? AND form_month=?",
            (y, m)
        ).fetchone()["next_seq"]

        db.execute("""
            INSERT INTO loans
            (created_at, start_date, duration_days, device_code,
             borrower_name, borrower_division, borrower_signature, borrower_signed_at, status,
             form_seq, form_month, form_year)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'submitted', ?, ?, ?)
        """, (now, now, ddays, device_code,
              borrower_name, borrower_division, signature_data, now,
              seq, m, y))

        db.execute("UPDATE devices SET status = 'reserved' WHERE code = ?", (device_code,))
        db.commit()

        return redirect(url_for("success", msg="Form peminjaman berhasil disubmit. Menunggu IT mengambil tiket."))

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return render_template("create_form.html", devices=devices, server_now=now, borrower_division=borrower_division_val)

# IT take ticket -> aktifkan
@app.route("/confirm-it/<int:loan_id>", methods=["GET", "POST"])
@login_required
def confirm_it(loan_id):
    if not session.get("is_it"):
        abort(403)
    db = get_db()
    row = db.execute("SELECT * FROM loans WHERE id = ?", (loan_id,)).fetchone()
    if not row:
        flash("Form tidak ditemukan.", "danger")
        return redirect(url_for("submitted_list"))
    if row["status"] != "submitted":
        flash("Form ini tidak dalam status 'submitted'.", "warning")
        return redirect(url_for("submitted_list"))

    # Mengambil kondisi terakhir barang dari peminjaman sebelumnya
    last_condition_row = db.execute("""
        SELECT return_status_kondisi, return_keterangan_kondisi
        FROM loans
        WHERE device_code = ? AND status = 'returned'
        ORDER BY return_it_confirmed_at DESC
        LIMIT 1
    """, (row["device_code"],)).fetchone()

    last_condition = {
        "status": "Belum pernah di cek sebelumnya",
        "keterangan": ""
    }
    if last_condition_row:
        last_condition["status"] = last_condition_row["return_status_kondisi"]
        last_condition["keterangan"] = last_condition_row["return_keterangan_kondisi"]

    if request.method == "POST":
        it_sig = request.form.get("signature_data", "")
        status_kondisi = request.form.get("status_kondisi", "").strip()
        keterangan_kondisi = request.form.get("keterangan_kondisi", "").strip()

        if not it_sig:
            flash("Tanda tangan PIC IT wajib.", "warning")
            return render_template("confirm_it.html",
                                   row=row,
                                   it_user=session.get("display_name") or session["username"],
                                   server_now=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                   last_condition=last_condition)

        if status_kondisi == "Rusak" and not keterangan_kondisi:
            flash("Keterangan kerusakan wajib diisi.", "warning")
            return render_template("confirm_it.html",
                                   row=row,
                                   it_user=session.get("display_name") or session["username"],
                                   server_now=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                   last_condition=last_condition)

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        db.execute("""
            UPDATE loans
            SET it_confirmed_by = ?, it_signature = ?, it_confirmed_at = ?, status = 'active', status_kondisi = ?, keterangan_kondisi = ?
            WHERE id = ?
        """, ((session.get("display_name") or session["username"]), it_sig, now, status_kondisi, keterangan_kondisi, loan_id))

        db.execute("UPDATE devices SET status = 'allocated' WHERE code = ?", (row["device_code"],))
        db.commit()
        return redirect(url_for("success", msg="Peminjaman diaktifkan."))

    return render_template("confirm_it.html",
                           row=row,
                           it_user=session.get("display_name") or session["username"],
                           server_now=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                           last_condition=last_condition)

# Daftar pinjaman aktif (untuk borrower memulai pengembalian)
@app.route("/return-active")
@login_required
def return_active_list():
    db = get_db()
    if session.get("is_it"):
        rows = db.execute(
            "SELECT * FROM loans WHERE status = 'active' ORDER BY created_at DESC"
        ).fetchall()
    else:
        # filter sesuai divisi user
        if not session.get("user_division"):
            return redirect(url_for("select_division"))
        rows = db.execute(
            "SELECT * FROM loans WHERE status='active' AND lower(borrower_division)=lower(?) ORDER BY created_at DESC",
            (session["user_division"],)
        ).fetchall()
    return render_template("return_active_list.html", rows=rows)

# Borrower isi form pengembalian -> return_submitted
@app.route("/return/<int:loan_id>", methods=["GET", "POST"])
@login_required
def return_form(loan_id):
    db = get_db()
    row = db.execute("SELECT * FROM loans WHERE id = ?", (loan_id,)).fetchone()
    if not row:
        flash("Form tidak ditemukan.", "danger")
        return redirect(url_for("return_active_list"))
    if row["status"] != "active":
        flash("Hanya dapat mengembalikan form dengan status aktif.", "warning")
        return redirect(url_for("return_active_list"))

    # Non-IT tidak boleh mengisi return untuk divisi lain
    if not session.get("is_it"):
        if not session.get("user_division"):
            return redirect(url_for("select_division"))
        if (row["borrower_division"] or "").lower() != session["user_division"].lower():
            abort(403)

    if request.method == "POST":
        ret_name = request.form.get("return_borrower_name", "").strip()
        ret_div  = request.form.get("return_division", "").strip()
        ret_sig  = request.form.get("signature_data", "")
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not all([ret_name, ret_div, ret_sig]):
            flash("Nama, divisi, dan tanda tangan pengembali wajib diisi.", "warning")
            return render_template("return_form.html", row=row, server_now=now, divisions=DIVISIONS)

        db.execute("""
            UPDATE loans
            SET return_borrower_name = ?, return_division = ?, return_signature = ?, return_submitted_at = ?, status = 'return_submitted'
            WHERE id = ?
        """, (ret_name, ret_div, ret_sig, now, loan_id))
        db.commit()
        return redirect(url_for("success", msg="Form pengembalian terkirim. Menunggu konfirmasi IT."))

    return render_template("return_form.html",
                           row=row,
                           server_now=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                           divisions=DIVISIONS)

# Queue forms pengembalian (IT)
@app.route("/returns")
@login_required
def returns_list():
    if not session.get("is_it"):
        abort(403)
    db = get_db()

    selected_division = request.args.get("divisi", "")

    all_divisions_rows = db.execute(
        "SELECT DISTINCT return_division FROM loans WHERE return_division IS NOT NULL AND return_division != '' ORDER BY return_division"
    ).fetchall()
    all_divisions = [row['return_division'] for row in all_divisions_rows]

    query = "SELECT * FROM loans WHERE status = 'return_submitted'"
    params = []

    if selected_division:
        query += " AND return_division = ?"
        params.append(selected_division)

    query += " ORDER BY return_submitted_at DESC"
    rows = db.execute(query, params).fetchall()

    return render_template(
        "returns_list.html",
        rows=rows,
        all_divisions=all_divisions,
        selected_division=selected_division
    )

# IT konfirmasi pengembalian -> returned
@app.route("/confirm-return-it/<int:loan_id>", methods=["GET", "POST"])
@login_required
def confirm_return_it(loan_id):
    if not session.get("is_it"):
        abort(403)
    db = get_db()
    row = db.execute("SELECT * FROM loans WHERE id = ?", (loan_id,)).fetchone()
    if not row:
        flash("Form tidak ditemukan.", "danger")
        return redirect(url_for("returns_list"))
    if row["status"] != "return_submitted":
        flash("Form ini tidak dalam status 'return_submitted'.", "warning")
        return redirect(url_for("returns_list"))

    if request.method == "POST":
        checked = request.form.get("checked_ack") == "on"
        it_sig  = request.form.get("signature_data", "")
        
        return_status_choice = request.form.get("return_status_kondisi", "").strip()
        new_keterangan = request.form.get("return_keterangan_kondisi", "").strip()

        final_status = ""
        final_keterangan = ""

        if return_status_choice == 'Rusak':
            final_status = 'Rusak'
            final_keterangan = new_keterangan
            if not final_keterangan:
                 flash("Keterangan kerusakan wajib diisi.", "warning")
                 return render_template("confirm_return_it.html",
                                        row=row,
                                        it_user=session.get("display_name") or session["username"],
                                        LEGAL_TEXT=LEGAL_TEXT)
        else: # 'Baik' (Sesuai kondisi saat peminjaman)
            final_status = row['status_kondisi'] or 'Baik'
            final_keterangan = row['keterangan_kondisi']

        if not checked:
            flash("Checkbox pernyataan wajib dicentang.", "warning")
            return render_template("confirm_return_it.html",
                                   row=row,
                                   it_user=session.get("display_name") or session["username"],
                                   LEGAL_TEXT=LEGAL_TEXT)
        if not it_sig:
            flash("Tanda tangan PIC IT wajib.", "warning")
            return render_template("confirm_return_it.html",
                                   row=row,
                                   it_user=session.get("display_name") or session["username"],
                                   LEGAL_TEXT=LEGAL_TEXT)

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        db.execute("""
            UPDATE loans
            SET return_it_confirmed_by = ?, return_it_checked_box = 1, return_it_confirmed_at = ?, return_it_signature = ?, status = 'returned',
                return_status_kondisi = ?, return_keterangan_kondisi = ?
            WHERE id = ?
        """, ((session.get("display_name") or session["username"]), now, it_sig, final_status, final_keterangan, loan_id))
        db.execute("UPDATE devices SET status = 'available' WHERE code = ?", (row["device_code"],))
        db.commit()
        return redirect(url_for("success", msg="Pengembalian dikonfirmasi. Device kembali available."))

    return render_template("confirm_return_it.html",
                           row=row,
                           it_user=session.get("display_name") or session["username"],
                           LEGAL_TEXT=LEGAL_TEXT)

# ======== HISTORY (IT only) ========
MONTHS_ID = [
    (1,"Januari"), (2,"Februari"), (3,"Maret"), (4,"April"), (5,"Mei"), (6,"Juni"),
    (7,"Juli"), (8,"Agustus"), (9,"September"), (10,"Oktober"), (11,"November"), (12,"Desember")
]

@app.route("/history")
@login_required
def history():
    if not session.get("is_it"):
        abort(403)
    db = get_db()
    now = datetime.now()
    bulan_q = request.args.get("bulan", type=str, default=str(now.month))
    tahun_q = request.args.get("tahun", type=str, default=str(now.year))
    applied  = request.args.get("applied") == "1"

    bulan_map = {str(i): i for i,_ in MONTHS_ID}
    bulan_map.update({name.lower(): i for i,name in MONTHS_ID})
    bln = bulan_map.get(bulan_q.lower() if bulan_q else "", now.month)
    thn = int(tahun_q) if (tahun_q and tahun_q.isdigit()) else now.year

    rows = db.execute("""
        SELECT * FROM loans
        WHERE strftime('%Y', created_at) = ? AND strftime('%m', created_at) = ?
        ORDER BY datetime(created_at) ASC
    """, (str(thn), f"{bln:02d}")).fetchall()

    return render_template("history.html", rows=rows, bulan=bln, tahun=thn, months=MONTHS_ID, applied=applied)

# ============ Export Excel (gambar centered H/V) ============
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string
from openpyxl.utils.units import pixels_to_EMU, points_to_pixels
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image as PILImage
from tempfile import NamedTemporaryFile

def image_from_data_url(data_url: str):
    if not data_url:
        return None
    m = re.match(r"data:image/(png|jpeg|jpg);base64,([\s\S]+)", data_url, re.I)
    if not m:
        return None
    try:
        raw = base64.b64decode(m.group(2))
        img = PILImage.open(io.BytesIO(raw)).convert("RGBA")
        return img
    except Exception:
        return None

def cm_to_px(cm: float, dpi: int = 96) -> int:
    return int(round((cm / 2.54) * dpi))

def cm_to_points(cm: float) -> float:
    return (cm / 2.54) * 72.0

def process_signature_image(data_url: str, w_cm: float = SIG_W_CM, h_cm: float = SIG_H_CM, crop: bool = SIG_CROP):
    """Resize/crop ttangan ke canvas putih w_cm x h_cm (cm)."""
    img = image_from_data_url(data_url)
    if img is None:
        return None
    target_w_px = cm_to_px(w_cm)
    target_h_px = cm_to_px(h_cm)
    target_ratio = target_w_px / target_h_px
    w, h = img.size
    ratio = w / h if h else 1.0

    if crop:
        # crop center ke rasio target
        if ratio > target_ratio:
            new_w = int(h * target_ratio)
            left = (w - new_w) // 2
            img = img.crop((left, 0, left + new_w, h))
        elif ratio < target_ratio:
            new_h = int(w / target_ratio)
            top = (h - new_h) // 2
            img = img.crop((0, top, w, top + new_h))
        img = img.resize((target_w_px, target_h_px), PILImage.LANCZOS)
        bg = PILImage.new("RGB", (target_w_px, target_h_px), (255, 255, 255))
        bg.paste(img, (0, 0), img)
        return bg
    else:
        # fit (tanpa crop), center dengan padding
        if ratio > target_ratio:
            new_w = target_w_px
            new_h = int(round(new_w / ratio))
        else:
            new_h = target_h_px
            new_w = int(round(new_h * ratio))
        img_resized = img.resize((new_w, new_h), PILImage.LANCZOS)
        bg = PILImage.new("RGB", (target_w_px, target_h_px), (255, 255, 255))
        x = (target_w_px - new_w) // 2
        y = (target_h_px - new_h) // 2
        bg.paste(img_resized, (x, y), img_resized)
        return bg

def save_pil_to_tempfile(img: PILImage.Image) -> str:
    tmp = NamedTemporaryFile(suffix=".png", delete=False)
    tmp.close()
    img.save(tmp.name, format="PNG")
    return tmp.name

def excel_col_width_to_pixels(width_chars: float | None) -> int:
    # Aproksimasi konversi
    if width_chars is None:
        width_chars = 8.43
    return int(round(width_chars * 7 + 5))

def center_image_in_cell(ws, xlimg: XLImage, cell_addr: str, img_w_px: int, img_h_px: int):
    """Letakkan gambar tepat di tengah cell (H&V)."""
    m = re.match(r"([A-Z]+)(\d+)", cell_addr, re.I)
    if not m:
        xlimg.anchor = cell_addr
        return
    col_letter = m.group(1).upper()
    row_idx = int(m.group(2))

    col_w_chars = (ws.column_dimensions[col_letter].width
                   if col_letter in ws.column_dimensions else None)
    cell_w_px = excel_col_width_to_pixels(col_w_chars)

    row_h_pt = (ws.row_dimensions[row_idx].height
                if row_idx in ws.row_dimensions else None)
    if row_h_pt is None:
        row_h_pt = 15  # default Excel (pt)
    cell_h_px = int(round(points_to_pixels(row_h_pt)))

    off_x_px = max(0, (cell_w_px - img_w_px) // 2)
    off_y_px = max(0, (cell_h_px - img_h_px) // 2)

    col_idx0 = column_index_from_string(col_letter) - 1
    marker = AnchorMarker(
        col=col_idx0, colOff=pixels_to_EMU(off_x_px),
        row=row_idx - 1, rowOff=pixels_to_EMU(off_y_px)
    )
    size = XDRPositiveSize2D(pixels_to_EMU(img_w_px), pixels_to_EMU(img_h_px))
    xlimg.anchor = OneCellAnchor(_from=marker, ext=size)

@app.route("/export-history")
@login_required
def export_history():
    if not session.get("is_it"):
        abort(403)

    now = datetime.now()
    bulan_q = request.args.get("bulan", type=str, default=str(now.month))
    tahun_q = request.args.get("tahun", type=str, default=str(now.year))

    bulan_map = {str(i): i for i,_ in MONTHS_ID}
    bulan_map.update({name.lower(): i for i,name in MONTHS_ID})
    bln = bulan_map.get(bulan_q.lower() if bulan_q else "", now.month)
    thn = int(tahun_q) if (tahun_q and tahun_q.isdigit()) else now.year

    db = get_db()
    rows = db.execute("""
        SELECT * FROM loans
        WHERE strftime('%Y', created_at) = ? AND strftime('%m', created_at) = ?
        ORDER BY datetime(created_at) ASC
    """, (str(thn), f"{bln:02d}")).fetchall()

    if not os.path.isfile(TEMPLATE_XLSX_PATH):
        abort(500, description=f"Template Excel tidak ditemukan: {TEMPLATE_XLSX_PATH}")

    wb = load_workbook(TEMPLATE_XLSX_PATH)
    ws = wb.active
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

    start_row = 9
    tmp_paths = []

    def add_sig(cell_addr: str, data_url: str):
        img_pil = process_signature_image(data_url)
        if img_pil is None:
            return
        w_px, h_px = img_pil.size
        path = save_pil_to_tempfile(img_pil)
        tmp_paths.append(path)
        xlimg = XLImage(path)
        center_image_in_cell(ws, xlimg, cell_addr, w_px, h_px)
        ws.add_image(xlimg)

    for idx, r in enumerate(rows, start=0):
        erow = start_row + idx

        # Tinggi baris diset dulu agar hitung vertikal akurat
        target_row_height_pt = cm_to_points(SIG_H_CM) + 8.0
        current_pt = ws.row_dimensions[erow].height if erow in ws.row_dimensions else None
        if current_pt is None or current_pt < target_row_height_pt:
            ws.row_dimensions[erow].height = target_row_height_pt

        ws[f"B{erow}"].value = idx + 1
        ws[f"C{erow}"].value = (r["created_at"].split(" ")[0] if r["created_at"] else "")
        ws[f"D{erow}"].value = r["device_code"]
        ws[f"E{erow}"].value = r["duration_days"]

        ws[f"F{erow}"].value = r["borrower_name"]
        ws[f"G{erow}"].value = r["borrower_division"]
        ws[f"I{erow}"].value = r["start_date"] or ""

        ws[f"J{erow}"].value = r["it_confirmed_by"] or ""

        ws[f"L{erow}"].value = r["return_borrower_name"] or ""
        ws[f"M{erow}"].value = r["return_division"] or ""
        ws[f"O{erow}"].value = r["return_submitted_at"] or ""

        ws[f"P{erow}"].value = r["return_it_confirmed_by"] or ""

        for col in "BCDEFGHIJKLMNOPQ":
            ws[f"{col}{erow}"].alignment = CENTER

        add_sig(f"H{erow}", r["borrower_signature"])
        add_sig(f"K{erow}", r["it_signature"])
        add_sig(f"N{erow}", r["return_signature"])
        add_sig(f"Q{erow}", r["return_it_signature"])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    for p in tmp_paths:
        try:
            os.remove(p)
        except Exception:
            pass

    filename = f"History_Peminjaman_{thn}-{bln:02d}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

# Alias lama
@app.route("/close/<int:loan_id>")
@login_required
def close_form_alias(loan_id):
    return redirect(url_for("return_form", loan_id=loan_id))

# Halaman success umum
@app.route("/success")
@login_required
def success():
    msg = request.args.get("msg", "Berhasil.")
    return render_template("success.html", message=msg)

# Healthcheck
@app.route("/healthz")
def healthz():
    return {"ok": True}, 200

# ===== Legal text (dipakai pada confirm_return_it) =====
LEGAL_TEXT = (
    "Saya telah memeriksa dan memastikan barang yang user kembalikan dalam keadaan baik, lengkap, "
    "berfungsi normal, dan sesuai daftar aset yang dipinjam."
)

# ===== Main =====
if __name__ == "__main__":
    host = os.getenv("HOST", "0.0.0.0")
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("DEBUG", "0") == "1"
    app.run(host=host, port=port, debug=debug)

